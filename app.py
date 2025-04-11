from flask import Flask, request, render_template, send_file
import pandas as pd
import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import io

app = Flask(__name__)

# ---- Your existing formatting functions ----
def add_time_headers(ws):
    previous_time = None
    max_rows = ws.max_row
    rows_to_insert = []

    for i in range(2, max_rows):
        cell_time = ws.cell(i, 2).value
        if cell_time != previous_time:
            rows_to_insert.append((i, cell_time))
        previous_time = cell_time

    for (r, time_val) in reversed(rows_to_insert):
        ws.insert_rows(r)
        cell = ws.cell(r, 1)
        cell.value = time_val
        cell.alignment = Alignment(horizontal="center", vertical="center")

def add_course_headers(ws):
    prev_course = None
    max_rows = ws.max_row
    rows_to_insert = []

    for i in range(2, max_rows):
        cell_course = ws.cell(i, 1).value
        cell_course_name = ws.cell(i, 4).value
        if cell_course != prev_course and ws.cell(i, 2).value is not None:
            rows_to_insert.append((i, cell_course_name))
            prev_course = cell_course

    for (r, header_val) in reversed(rows_to_insert):
        ws.insert_rows(r)
        cell = ws.cell(r, 1)
        cell.value = header_val
        #cell.alignment = Alignment(horizontal="center", vertical="center")

def add_instructor_headers(ws):
    max_rows = ws.max_row
    for i in range(2, max_rows):
        cell_instr = ws.cell(i + 1, 3).value
        if ws.cell(i, 2).value is None and ':' not in str(ws.cell(i, 1).value):
            ws.cell(i, 1).value = f"{ws.cell(i, 1).value} - {cell_instr}"

# ---- End of formatting functions ----

def process_csv(df, options, instructors_classes):
    # Build a mapping of classes to instructors from the user input:
    print("pee")
    class_to_instructor = {}
    for instructor, classes in instructors_classes.items():
        for clas in classes:
            class_to_instructor[clas] = instructor

    # Process the CSV file
    columns_to_keep = ['EventID', 'EventTime', 'Service', 'AttendeeName', 'Phone']
    df_subset = df[columns_to_keep].copy()
    df_subset['Instructor'] = df_subset['EventID'].map(class_to_instructor)
    df_subset = df_subset[['EventID', 'EventTime', 'Instructor', 'Service', 'AttendeeName', 'Phone']]

    # Save DataFrame to Excel in memory
    output = io.BytesIO()
    df_subset.to_excel(output, index=False, sheet_name='Sheet1')
    output.seek(0)

    wb = load_workbook(output)
    ws1 = wb.active

    # Set default font size for each cell
    for row in ws1.iter_rows():
        for cell in row:
            cell.font = Font(size=12)

    # Apply formatting options based on user choices:
    if options.get('time_headers'):
        add_time_headers(ws1)
    if options.get('course_headers'):
        add_course_headers(ws1)
    if options.get('instructor_headers'):
        add_instructor_headers(ws1)

    # Example additional formatting:
    max_rows = ws1.max_row
    if options.get('bold_time'):
        for i in range(2, max_rows):
            if ws1.cell(i, 2).value is None and ':' in str(ws1.cell(i, 1).value):
                ws1.cell(i, 1).font = Font(bold=True)
                ws1.row_dimensions[i].height = 20

    if options.get('bold_course'):
        for i in range(2, max_rows):
            if ws1.cell(i, 2).value is None and ':' not in ws1.cell(i, 1).value:
                ws1.cell(i, 1).font = Font(bold=True)
                ws1.row_dimensions[i].height = 20

    if options.get('center_time'):
        for i in range(2, max_rows):
            if ws1.cell(i, 2).value is None and ':' in str(ws1.cell(i, 1).value):
                ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
                ws1.row_dimensions[i].height = 20
    
    if options.get('center_course'):
        for i in range(2, max_rows):
            if ws1.cell(i, 2).value is None and ':' not in ws1.cell(i, 1).value:
                ws1.merge_cells(start_row=i, start_column=1, end_row=i,
                                end_column=7)
                ws1.row_dimensions[i].height = 20
            
    if options.get('borders'):
        print("poop")
        for i in range(2, max_rows + 1):
            if ws1.cell(i, 2).value is None:

                ws1.cell(i, 1).border = Border(left=Side(style='thin'),
                                               right=Side(style='thin'),
                                               top=Side(style='thin'),
                                               bottom=Side(style='thin'))
                ws1.cell(i, 2).border = Border(top=Side(style='thin'),
                                               bottom=Side(style='thin'))
                ws1.cell(i, 3).border = Border(top=Side(style='thin'),
                                                  bottom=Side(style='thin'))
                ws1.cell(i, 4).border = Border(top=Side(style='thin'),
                                                  bottom=Side(style='thin'))
                ws1.cell(i, 5).border = Border(top=Side(style='thin'),
                                                    bottom=Side(style='thin'))
                ws1.cell(i, 6).border = Border(top=Side(style='thin'),
                                                    bottom=Side(style='thin'))
                ws1.cell(i, 7).border = Border(top=Side(style='thin'),
                                                  bottom=Side(style='thin'),
                                                  right=Side(style='thin'))
            else:

                ws1.cell(i, 1).border = Border(left=Side(style='thin'))
                ws1.cell(i, 7).border = Border(right=Side(style='thin'))

            if i == max_rows:
                ws1.cell(i, 1).border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
                ws1.cell(i, 2).border = Border(bottom=Side(style='thin'))
                ws1.cell(i, 3).border = Border(bottom=Side(style='thin'))
                ws1.cell(i, 4).border = Border(bottom=Side(style='thin'))
                ws1.cell(i, 5).border = Border(bottom=Side(style='thin'))
                ws1.cell(i, 6).border = Border(bottom=Side(style='thin'))
                ws1.cell(i, 7).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    # Resize columns based on content
    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws1.column_dimensions[col_letter].width = max_length + 2
    ws1.column_dimensions['A'].width = 10

    # Save the workbook to a new BytesIO stream for download:
    output_final = io.BytesIO()
    now = dt.datetime.now()
    filename = f"MasterList_{now.month}_{now.day}_{now.year}.xlsx"
    wb.save(output_final)
    output_final.seek(0)
    return output_final, filename

@app.route('/', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        # Get the CSV file
        file = request.files.get('csv_file')
        if not file:
            return "No file uploaded", 400
        try:
            df = pd.read_csv(file)
        except Exception as e:
            return f"Error reading CSV: {e}", 400

        # Get formatting options (checkboxes)
        options = {
            'time_headers': 'time_headers' in request.form,
            'course_headers': 'course_headers' in request.form,
            'instructor_headers': 'instructor_headers' in request.form,
            'borders': 'borders' in request.form,
            'center_time': 'center_time' in request.form,
            'center_course': 'center_course' in request.form,
            'bold_time': 'bold_time' in request.form,
            'bold_course': 'bold_course' in request.form,
        }

        # ---- Process instructors input ----
        instructor_names = request.form.getlist('instructor_names[]')
        instructor_codes = request.form.getlist('instructor_codes[]')

        # Build the instructors_classes dictionary; convert each class code to integer if needed
        instructors_classes = {}
        for name, codes in zip(instructor_names, instructor_codes):
            if name.strip():
                # Convert comma-separated codes to a list; if you prefer strings, you can skip int conversion
                code_list = [int(code.strip()) for code in codes.split(',') if code.strip()]
                instructors_classes[name.strip()] = code_list

        # Process the CSV file using provided options and instructors data
        output_file, filename = process_csv(df, options, instructors_classes)

        return send_file(output_file,
                         as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('upload.html')

if __name__ == '__main__':
    app.run()
